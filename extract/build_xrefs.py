"""Build the Cross Reference Search index from DATA_Xrefs.xlsx.

Sheet layout: column A = our Part Number, column B = Description,
columns C..Q = one competitor brand per column (brand name in row 1),
its reference in that row when we have a match.

Writes public/data/xref-index.json as a flat array so the client can
search it directly, without a nav tree (this index is a standalone
lookup tool, separate from the Product Selector guide).
"""
import json
import openpyxl

SRC = "DATA_Xrefs.xlsx"
OUT = "public/data/xref-index.json"

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb.active
headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
brands = headers[2:]

entries = []
for r in range(2, ws.max_row + 1):
    pn = ws.cell(row=r, column=1).value
    desc = ws.cell(row=r, column=2).value
    if pn is None and desc is None:
        continue
    pn = str(pn).strip() if pn is not None else ""
    desc = str(desc).strip() if desc is not None else ""
    if not pn:
        continue
    refs = []
    for i, c in enumerate(range(3, 3 + len(brands))):
        v = ws.cell(row=r, column=c).value
        if v is None:
            continue
        v = str(v).strip()
        if not v:
            continue
        refs.append({"brand": brands[i], "ref": v})
    entries.append({"pn": pn, "desc": desc, "refs": refs})

with open(OUT, "w", encoding="utf-8") as f:
    json.dump(entries, f, ensure_ascii=False)

print("entries:", len(entries))
print("brands:", brands)
