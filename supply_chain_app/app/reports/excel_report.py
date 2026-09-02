"""Multi-sheet Excel report export (openpyxl), 100% local, no web dependency."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F3B57", end_color="1F3B57", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=16, bold=True, color="1F3B57")


def _write_df(ws, df: pd.DataFrame, start_row: int = 1) -> int:
    if df.empty:
        ws.cell(row=start_row, column=1, value="No data available.")
        return start_row + 1

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=str(col_name).replace("_", " ").title())
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            if isinstance(value, float):
                value = round(value, 2)
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, col_name in enumerate(df.columns, start=1):
        width = max(12, min(40, len(str(col_name)) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    return start_row + len(df) + 2


def build_excel_report(
    output_path: str | Path,
    kpis: dict,
    top_products: pd.DataFrame,
    abc_xyz: pd.DataFrame,
    coverage: pd.DataFrame,
    recommendations: pd.DataFrame,
) -> Path:
    wb = Workbook()

    ws = wb.active
    ws.title = "Dashboard"
    ws["A1"] = "Supply Chain & Sales Decision Assistant — Report"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    kpi_df = pd.DataFrame(
        [{"Metric": k.replace("_", " ").title(), "Value": v} for k, v in kpis.items()]
    )
    _write_df(ws, kpi_df, start_row=4)
    row = 4 + len(kpi_df) + 3
    ws.cell(row=row, column=1, value="Top Products by Revenue").font = Font(bold=True, size=12)
    _write_df(ws, top_products, start_row=row + 1)

    ws2 = wb.create_sheet("ABC-XYZ Analysis")
    _write_df(ws2, abc_xyz)

    ws3 = wb.create_sheet("Stock Coverage")
    _write_df(ws3, coverage)

    ws4 = wb.create_sheet("Recommendations")
    _write_df(ws4, recommendations)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
